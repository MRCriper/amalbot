import asyncio
import sqlite3
import secrets
import io
import os
from openpyxl import load_workbook
from PIL import Image, ImageDraw, ImageFont
from aiogram import Bot, Dispatcher, types, F
from aiogram.enums import ParseMode
from aiogram.filters import Command
from aiogram.client.default import DefaultBotProperties
from aiogram.fsm.context import FSMContext
from aiogram.fsm.state import State, StatesGroup

TOKEN = "7659805201:AAG0gtQM3P9U7aJH1BGkBm-Sdzm2H--zwY4"
ADMIN_ID = 1312455951

bot = Bot(
    token=TOKEN,
    default=DefaultBotProperties(parse_mode=ParseMode.HTML)
)
dp = Dispatcher()

# Состояния для FSM
class AuthStates(StatesGroup):
    waiting_code = State()
    waiting_last_name = State()
    waiting_first_name = State()

class AdminStates(StatesGroup):
    waiting_excel = State()

# Функция загрузки шрифта с поддержкой UTF-8
def get_unicode_font(size=14):
    """Загружает шрифт с поддержкой Unicode"""
    font_paths = [
        'DejaVuSans.ttf',
        'arial.ttf',
        'ARIAL.TTF',
        'NotoSans-Regular.ttf',
        '/System/Library/Fonts/Arial.ttf',  # macOS
        '/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf',  # Linux
        'C:/Windows/Fonts/arial.ttf',  # Windows
        'C:/Windows/Fonts/calibri.ttf',  # Windows
    ]
    
    for font_path in font_paths:
        try:
            return ImageFont.truetype(font_path, size)
        except:
            continue
    
    return ImageFont.load_default()

# Инициализация базы данных
def init_db():
    conn = sqlite3.connect('bot_data.db')
    cursor = conn.cursor()
    
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS access_codes (
            id INTEGER PRIMARY KEY,
            code TEXT UNIQUE,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    ''')
    
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS authorized_users (
            user_id INTEGER PRIMARY KEY,
            username TEXT,
            last_name TEXT,
            first_name TEXT,
            tabel_data TEXT,
            authorized_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    ''')
    
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS excel_data (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            fio TEXT,
            row_text TEXT
        )
    ''')
    
    conn.commit()
    conn.close()

# Проверка авторизации пользователя
def is_authorized(user_id):
    if user_id == ADMIN_ID:
        return True
    
    conn = sqlite3.connect('bot_data.db')
    cursor = conn.cursor()
    cursor.execute('SELECT user_id FROM authorized_users WHERE user_id = ?', (user_id,))
    result = cursor.fetchone()
    conn.close()
    return result is not None

# Генерация кода доступа (удаляем старый)
def generate_access_code():
    code = secrets.token_hex(4).upper()
    
    conn = sqlite3.connect('bot_data.db')
    cursor = conn.cursor()
    # Удаляем все старые коды
    cursor.execute('DELETE FROM access_codes')
    # Создаем новый код
    cursor.execute('INSERT INTO access_codes (code) VALUES (?)', (code,))
    conn.commit()
    conn.close()
    
    return code

# Проверка кода доступа (код работает многократно)
def verify_access_code(code):
    conn = sqlite3.connect('bot_data.db')
    cursor = conn.cursor()
    cursor.execute('SELECT id FROM access_codes WHERE code = ?', (code,))
    result = cursor.fetchone()
    conn.close()
    return result is not None

# ИСПРАВЛЕНО: Обработка Excel с сохранением пустых ячеек
def process_excel_file(file_path):
    try:
        wb = load_workbook(filename=file_path, data_only=True)
        
        # Ищем лист "СВОД (по месяцам)"
        if 'СВОД (по месяцам)' not in wb.sheetnames:
            return False, "Лист 'СВОД (по месяцам)' не найден в файле"
        
        ws = wb['СВОД (по месяцам)']
        
        conn = sqlite3.connect('bot_data.db')
        cursor = conn.cursor()
        
        # Удаляем старые данные
        cursor.execute('DELETE FROM excel_data')
        
        # Сохраняем заголовок (строка 4, столбцы A-N)
        header_values = []
        for col_idx in range(1, 15):  # A=1, N=14
            cell_value = ws.cell(row=4, column=col_idx).value
            # ВАЖНО: Сохраняем пустые ячейки как пустую строку, а не пропускаем
            if cell_value is not None:
                header_values.append(str(cell_value).strip())
            else:
                header_values.append("")  # Пустая ячейка остается пустой
        
        header_text = '|'.join(header_values)
        cursor.execute('INSERT INTO excel_data (fio, row_text) VALUES (?, ?)', ('_header', header_text))
        
        # Обрабатываем строки с 5 по 31
        for row_idx in range(5, 32):
            # ФИО из столбца B (индекс 2)
            fio_cell = ws.cell(row=row_idx, column=2)
            fio = str(fio_cell.value or '').strip()
            
            if not fio:  # Пропускаем пустые строки
                continue
                
            # Берем столбцы A-N (индексы 1-14)
            row_values = []
            for col_idx in range(1, 15):
                cell_value = ws.cell(row=row_idx, column=col_idx).value
                # ВАЖНО: Сохраняем пустые ячейки
                if cell_value is not None:
                    row_values.append(str(cell_value).strip())
                else:
                    row_values.append("")  # Пустая ячейка
            
            row_text = '|'.join(row_values)
            cursor.execute('INSERT INTO excel_data (fio, row_text) VALUES (?, ?)', (fio, row_text))
        
        conn.commit()
        conn.close()
        return True, "Файл успешно обработан"
        
    except Exception as e:
        return False, f"Ошибка обработки: {str(e)}"
    
# Получение заголовка
def get_header():
    conn = sqlite3.connect('bot_data.db')
    cursor = conn.cursor()
    cursor.execute('SELECT row_text FROM excel_data WHERE fio = ?', ('_header',))
    result = cursor.fetchone()
    conn.close()
    return result[0] if result else None

# Поиск строки пользователя по ФИО
def get_user_row_by_fio(last_name, first_name):
    full_name = f"{last_name} {first_name}"
    
    conn = sqlite3.connect('bot_data.db')
    cursor = conn.cursor()
    
    # Ищем точное совпадение
    cursor.execute('SELECT row_text FROM excel_data WHERE fio LIKE ? AND fio != ?', (f'%{full_name}%', '_header'))
    result = cursor.fetchone()
    
    # Если не найдено, ищем по фамилии
    if not result:
        cursor.execute('SELECT row_text FROM excel_data WHERE fio LIKE ? AND fio != ?', (f'{last_name}%', '_header'))
        result = cursor.fetchone()
    
    conn.close()
    return result[0] if result else None

# ИСПРАВЛЕНО: Правильное отображение пустых ячеек
def create_combined_image(header_text, row_text):
    if not header_text or not row_text:
        return None
    
    header_columns = header_text.split('|')
    row_columns = row_text.split('|')
    
    # Настройки изображения
    width = 1600
    row_height = 80
    height = row_height * 2 + 40
    background_color = (255, 255, 255)
    text_color = (0, 0, 0)
    header_bg_color = (240, 240, 240)
    border_color = (150, 150, 150)
    empty_cell_color = (250, 200, 200)  # Светло-красный фон для пустых ячеек
    
    # Создаем изображение
    image = Image.new('RGB', (width, height), background_color)
    draw = ImageDraw.Draw(image)
    
    font = get_unicode_font(16)
    col_width = width // max(len(header_columns), len(row_columns))
    
    # Функция для отрисовки строки
    def draw_row(y, columns, bg_color=None):
        if bg_color:
            draw.rectangle([0, y-15, width, y+row_height-15], fill=bg_color)
        
        for i, col_data in enumerate(columns):
            x = i * col_width + 10
            cell_x_start = i * col_width
            cell_x_end = (i + 1) * col_width
            
            # Проверяем, пустая ли ячейка
            is_empty = not str(col_data).strip()
            
            # Если ячейка пустая, закрашиваем ее особым цветом
            if is_empty and y > 50:  # Только для данных пользователя, не для заголовка
                draw.rectangle([cell_x_start, y-15, cell_x_end, y+row_height-15], 
                              fill=empty_cell_color)
                text = "—"  # Тире для пустых ячеек
            else:
                text = str(col_data).strip() if str(col_data).strip() else ""
            
            # Обрезаем длинный текст
            if len(text) > 15:
                text = text[:12] + "..."
            
            # Рисуем текст
            draw.text((x, y), text, fill=text_color, font=font)
            
            # Вертикальные линии между столбцами
            if i < len(columns) - 1:
                line_x = (i + 1) * col_width
                draw.line([(line_x, y-15), (line_x, y+row_height-15)], fill=border_color, width=1)
    
    # Рисуем заголовок
    draw_row(25, header_columns, header_bg_color)
    
    # Рисуем данные пользователя
    draw_row(105, row_columns)
    
    # Внешняя рамка
    draw.rectangle([0, 0, width-1, height-1], outline=border_color, width=2)
    
    # Горизонтальная линия между заголовком и данными
    draw.line([(0, 90), (width, 90)], fill=border_color, width=2)
    
    buffer = io.BytesIO()
    image.save(buffer, format='PNG')
    buffer.seek(0)
    return buffer

# Авторизация пользователя с именем и фамилией
def authorize_user_with_name(user_id, username, last_name, first_name):
    conn = sqlite3.connect('bot_data.db')
    cursor = conn.cursor()
    cursor.execute('''
        INSERT OR REPLACE INTO authorized_users 
        (user_id, username, last_name, first_name) 
        VALUES (?, ?, ?, ?)
    ''', (user_id, username, last_name, first_name))
    conn.commit()
    conn.close()

# Получение данных пользователя
def get_user_data(user_id):
    conn = sqlite3.connect('bot_data.db')
    cursor = conn.cursor()
    cursor.execute('SELECT last_name, first_name FROM authorized_users WHERE user_id = ?', (user_id,))
    result = cursor.fetchone()
    conn.close()
    return result if result else None

@dp.message(Command("start"))
async def start(message: types.Message, state: FSMContext):
    if is_authorized(message.from_user.id):
        if message.from_user.id == ADMIN_ID:
            # Админские кнопки
            keyboard = types.ReplyKeyboardMarkup(
                keyboard=[
                    [types.KeyboardButton(text="Загрузить табель Excel")],
                    [types.KeyboardButton(text="Сгенерировать код для входа")],
                    [types.KeyboardButton(text="Табель баллов"), types.KeyboardButton(text="Расписание звонков")]
                ],
                resize_keyboard=True
            )
            await message.answer("👑 Добро пожаловать, администратор!", reply_markup=keyboard)
        else:
            # Обычные пользователи - только одна кнопка
            keyboard = types.ReplyKeyboardMarkup(
                keyboard=[
                    [types.KeyboardButton(text="Табель баллов"), types.KeyboardButton(text="Расписание звонков")]
                ],
                resize_keyboard=True
            )
            await message.answer("👋 Добро пожаловать! Выберите нужный раздел:", reply_markup=keyboard)
    else:
        await message.answer("🔐 Для доступа к боту введите код:")
        await state.set_state(AuthStates.waiting_code)

@dp.message(AuthStates.waiting_code)
async def process_access_code(message: types.Message, state: FSMContext):
    code = message.text.upper().strip()
    
    if verify_access_code(code):
        await message.answer("✅ Код принят! Введите вашу фамилию:")
        await state.set_state(AuthStates.waiting_last_name)
    else:
        await message.answer("❌ Неверный код. Попробуйте еще раз:")

@dp.message(AuthStates.waiting_last_name)
async def process_last_name(message: types.Message, state: FSMContext):
    last_name = message.text.strip()
    await state.update_data(last_name=last_name)
    await message.answer("📝 Теперь введите ваше имя:")
    await state.set_state(AuthStates.waiting_first_name)

@dp.message(AuthStates.waiting_first_name)
async def process_first_name(message: types.Message, state: FSMContext):
    first_name = message.text.strip()
    user_data = await state.get_data()
    last_name = user_data['last_name']
    
    # Авторизуем пользователя с полными данными
    authorize_user_with_name(
        message.from_user.id, 
        message.from_user.username,
        last_name,
        first_name
    )
    
    await state.clear()
    
    keyboard = types.ReplyKeyboardMarkup(
        keyboard=[
            [types.KeyboardButton(text="Табель баллов"), types.KeyboardButton(text="Расписание звонков")]
        ],
        resize_keyboard=True
    )
    await message.answer(
        f"✅ Регистрация завершена!\nДобро пожаловать, {first_name} {last_name}!", 
        reply_markup=keyboard
    )

@dp.message(F.text == "Загрузить табель")
async def request_image_upload(message: types.Message, state: FSMContext):
    if message.from_user.id != ADMIN_ID:
        await message.answer("❌ Недостаточно прав")
        return
    
    await message.answer("📊 Отправьте Excel файл с табелем:")
    await state.set_state(AdminStates.waiting_excel)

@dp.message(AdminStates.waiting_excel, F.document)
async def handle_excel_upload(message: types.Message, state: FSMContext):
    if message.from_user.id != ADMIN_ID:
        await message.answer("❌ Недостаточно прав")
        return
    
    try:
        # Скачиваем файл
        file = await bot.get_file(message.document.file_id)
        file_path = f"temp_{message.document.file_id}.xlsx"
        await bot.download_file(file.file_path, file_path)
        
        # Обрабатываем Excel
        success, message_text = process_excel_file(file_path)
        
        if success:
            await message.answer(f"✅ {message_text}")
        else:
            await message.answer(f"❌ {message_text}")
        
        # Удаляем временный файл
        if os.path.exists(file_path):
            os.remove(file_path)
            
    except Exception as e:
        await message.answer(f"❌ Ошибка: {str(e)}")
    
    await state.clear()

@dp.message(F.text == "Сгенерировать код для входа")
async def generate_code(message: types.Message):
    if message.from_user.id != ADMIN_ID:
        await message.answer("❌ Недостаточно прав")
        return
    
    code = generate_access_code()
    await message.answer(f"🔑 Новый код доступа: <code>{code}</code>\n\n⚠️ Старый код удален. Этот код можно использовать многократно.")

@dp.message(F.text == "Табель баллов")
async def show_grades(message: types.Message):
    if not is_authorized(message.from_user.id):
        await message.answer("❌ Нет доступа к боту")
        return
    
    if message.from_user.id == ADMIN_ID:
        await message.answer("👑 Как администратор, вы видите весь табель. Обычные пользователи видят только свою строку.")
        return
    
    user_data = get_user_data(message.from_user.id)
    if not user_data:
        await message.answer("❌ Ваши данные не найдены. Обратитесь к администратору.")
        return
    
    last_name, first_name = user_data

     # Получаем заголовок и строку пользователя
    header_text = get_header()
    row_text = get_user_row_by_fio(last_name, first_name)
    
    if not header_text:
        await message.answer("❌ Заголовок табеля не найден. Обратитесь к администратору.")
        return
    
    row_text = get_user_row_by_fio(last_name, first_name)
    
    if not row_text:
        await message.answer(f"❌ Данные для {first_name} {last_name} не найдены в табеле.")
        return
    
    # Создаем комбинированное изображение
    image_buffer = create_combined_image(header_text, row_text)
    if image_buffer:
        from aiogram.types import BufferedInputFile
        
        photo = BufferedInputFile(
            image_buffer.getvalue(),
            filename=f"tabel_{first_name}_{last_name}.png"
        )
        
        await message.answer_photo(
            photo=photo,
            caption=f"📊 Табель баллов для {first_name} {last_name}"
        )
    else:
        await message.answer("❌ Ошибка при создании изображения")

# НОВАЯ ФУНКЦИЯ: Отправка расписания звонков
@dp.message(F.text == "Расписание звонков")
async def show_bell_schedule(message: types.Message):
    if not is_authorized(message.from_user.id):
        await message.answer("❌ Нет доступа к боту")
        return
    
    try:
        # Проверяем наличие файла bell.jpg
        if os.path.exists("bell.jpg"):
            from aiogram.types import FSInputFile
            
            photo = FSInputFile("bell.jpg")
            await message.answer_photo(
                photo=photo,
                caption="🔔 Расписание звонков"
            )
        else:
            await message.answer("❌ Файл с расписанием звонков не найден. Обратитесь к администратору.")
    
    except Exception as e:
        await message.answer(f"❌ Ошибка при отправке расписания: {str(e)}")

@dp.message()
async def handle_unauthorized(message: types.Message):
    if not is_authorized(message.from_user.id):
        await message.answer("❌ Для доступа к боту введите /start")

if __name__ == "__main__":
    init_db()
    asyncio.run(dp.start_polling(bot))
